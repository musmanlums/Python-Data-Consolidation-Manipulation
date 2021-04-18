# -*- coding: utf-8 -*-
"""
Created on Mon Apr 12 09:05:15 2021
"""
import os  # OS = OPERATING SYSTEM

os.chdir('D:/For_consolidation')  # TELLS PYTHON TO GO TO THE FOLDER BREVILLE AUTOMATION

import pandas as pd  # PANDAS
from openpyxl import load_workbook  # PYTHON EXCEL LIBRARY, VERY USEFUL
from datetime import datetime  # RECORD HOW LONG THIS TAKES

# from packages import SQLProjectMgmt

start = datetime.now()

# PASS A LSIT ON WHAT I NEED TO GET: BU - OPERATING UNITS
bu = ['AUO', 'AUD', 'CNP', 'UST', 'DET', 'UKT', 'HKD']

# LOOP OVER WORKBOOKS
bu_info = []  # EVERYTHING I SCRAPE IS STORED IN THIS OBJECT

for file in os.listdir('D:/For_consolidation'):

    wb = load_workbook(filename=file, data_only=True)
    print('file:', file, 'loaded.')

    for sheet in wb.sheetnames:
        if sheet in bu:
            ws = wb[sheet]
            costcentre = [ws.cell(column=1, row=i).value for i in range(2, 252)]
            accounts = [ws.cell(column=6, row=i).value for i in range(2, 252)]
            mar21 = [ws.cell(column=16, row=i).value for i in range(2, 252)]
            apr21 = [ws.cell(column=17, row=i).value for i in range(2, 252)]
            may21 = [ws.cell(column=18, row=i).value for i in range(2, 252)]
            jun21 = [ws.cell(column=19, row=i).value for i in range(2, 252)]

            d = {'Cost Centre': costcentre, 'Accounts': accounts, 'Mar 21': mar21, 'April 21': apr21, 'May 21': may21,
                 'Jun 21': jun21}
            df = pd.DataFrame(d)
            df['BusinessUnit'] = str(sheet)
            bu_info.append(df)

# list of dataframes, if all the dataframes have the same structure, pd.concat creates on big dataframe for all
bu_info = pd.concat(bu_info)
bu_info.to_excel('Consolidated.xlsx', index=False)

fileName = 'Consolidated.xlsx'
df = pd.read_excel(fileName)
print(df)

df_1 = pd.melt(df, id_vars=['Cost Centre', 'Business Unit', 'Accounts'], value_vars=['Mar 21', 'April 21', 'May 21', 'Jun 21'])
print(df_1)

df_2 = df_1.sort_values(['Cost Centre', 'Business Unit', 'Accounts'])
df_2.to_excel('Consolidated_pivoted.xlsx', index=False)